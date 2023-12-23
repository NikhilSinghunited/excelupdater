from flask import Flask, render_template, request
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/convert', methods=['POST'])
def convert():
    input_text = request.form['input_text']
    
    # Process the input_text as needed
    output_text = f"{input_text}"  # Assuming you want to include 'Input: ' before the actual input

    # Save data to Excel file
    excel_file_path = 'Databaseviewer.xlsx'
    try:
        # Load the workbook
        workbook = load_workbook(filename=excel_file_path)

        # Select the active sheet
        sheet = workbook.active

        # Access the input and output columns using the get_column_letter function
        input_column_letter = get_column_letter(1)  # Column A is 1
        output_column_letter = get_column_letter(2)  # Column B is 2

        # Find the last used row in either column (assuming both columns have the same length)
        last_row = max(sheet.max_row, sheet.max_column)

        # Insert 'Kishan' in the input column and 'Tiwari' in the output column at the next available row
        sheet[input_column_letter + str(last_row + 1)] = input_text
        sheet[output_column_letter + str(last_row + 1)] = output_text

        # Save the modified workbook
        workbook.save(filename=excel_file_path)

        return render_template('index.html', output=output_text, success=True)
    except Exception as e:
        return render_template('index.html', output=f"Error: {e}", success=False)

if __name__ == '__main__':
    app.run(debug=True)
